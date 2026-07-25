"""Export service — generates Excel (XLSX) files for reports"""

import io
from datetime import date
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers


HEADER_FONT = Font(name="Inter", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
TITLE_FONT = Font(name="Inter", bold=True, size=14, color="1E40AF")
SUBTITLE_FONT = Font(name="Inter", size=10, color="6B7280")
DATA_FONT = Font(name="Inter", size=10)
BOLD_FONT = Font(name="Inter", bold=True, size=10)
CURRENCY_FMT = '#,##0'
DATE_FMT = "DD/MM/YYYY"
THIN_BORDER = Border(
    bottom=Side(style="thin", color="E5E7EB"),
)


def _style_header(ws, cols):
    for col_idx in range(1, cols + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _write_title(ws, titulo, fecha_desde, fecha_hasta):
    ws.cell(row=1, column=1, value=titulo).font = TITLE_FONT
    ws.cell(row=2, column=1).font = SUBTITLE_FONT
    periodo = f"Período: {fecha_desde or 'Inicio'} — {fecha_hasta or 'Actual'}" if fecha_desde or fecha_hasta else "Todos los períodos"
    ws.cell(row=2, column=1, value=periodo)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)


def _write_data(ws, headers, rows, start_row=3):
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=start_row, column=col_idx, value=header)
    _style_header(ws, len(headers))
    for row_idx, row_data in enumerate(rows, start_row + 1):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if isinstance(val, (int, float)) and col_idx > 2:
                cell.number_format = CURRENCY_FMT


def _auto_width(ws):
    from openpyxl.utils import get_column_letter
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            idx = getattr(cell, "column", None)
            if not idx:  # celdas combinadas sin índice → saltar
                continue
            widths[idx] = max(widths.get(idx, 0), len(str(cell.value)))
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(w + 4, 35)


def export_sales_summary(data: dict, fecha_desde: Optional[date] = None, fecha_hasta: Optional[date] = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Ventas"

    _write_title(ws, "Resumen de Ventas", fecha_desde, fecha_hasta)

    headers = ["Indicador", "Valor"]
    rows = [
        ("Total de ventas", data.get("total_ventas", 0)),
        ("Monto total", data.get("monto_total", 0)),
        ("Monto IVA 10%", data.get("monto_iva_10", 0)),
        ("Monto IVA 5%", data.get("monto_iva_5", 0)),
        ("Monto exento", data.get("monto_exento", 0)),
        ("Ticket promedio", data.get("ticket_promedio", 0)),
        ("Total items vendidos", data.get("total_items", 0)),
    ]
    _write_data(ws, headers, rows)

    for r in range(4, 4 + len(rows)):
        ws.cell(row=r, column=1).font = BOLD_FONT
        ws.cell(row=r, column=2).number_format = CURRENCY_FMT

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_sales_by_period(data: list, fecha_desde: Optional[date] = None, fecha_hasta: Optional[date] = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas por Período"

    _write_title(ws, "Ventas por Período", fecha_desde, fecha_hasta)

    headers = ["Período", "Cantidad", "Monto Total", "IVA 10%", "Items"]
    rows = [(r["periodo"], r["cantidad"], r["monto"], r["iva_10"], r["items"]) for r in data]
    _write_data(ws, headers, rows)

    ws.cell(row=4, column=1).font = BOLD_FONT

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_sales_by_category(data: list, fecha_desde: Optional[date] = None, fecha_hasta: Optional[date] = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas por Categoría"

    _write_title(ws, "Ventas por Categoría", fecha_desde, fecha_hasta)

    headers = ["Categoría", "Cantidad", "Monto Total", "% Participación"]
    rows = [(r["categoria"], r["cantidad"], r["monto"], f"{r['porcentaje']}%") for r in data]
    _write_data(ws, headers, rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_sales_by_product(data: list, fecha_desde: Optional[date] = None, fecha_hasta: Optional[date] = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Top Productos"

    _write_title(ws, "Top Productos Vendidos", fecha_desde, fecha_hasta)

    headers = ["Producto", "SKU", "Cantidad", "Monto Total", "Costo", "Margen %"]
    rows = [(r["producto"], r["sku"], r["cantidad"], r["monto"], r["costo"], f"{r['margen']}%") for r in data]
    _write_data(ws, headers, rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_sales_by_client(data: list, fecha_desde: Optional[date] = None, fecha_hasta: Optional[date] = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas por Cliente"

    _write_title(ws, "Ventas por Cliente", fecha_desde, fecha_hasta)

    headers = ["Cliente", "RUC", "Cant. Compras", "Monto Total", "Última Compra"]
    rows = [(r["cliente"], r["ruc"], r["cantidad_compras"], r["monto_total"], r["ultima_compra"]) for r in data]
    _write_data(ws, headers, rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_inventory_summary(summary: dict, detail: list, fecha_desde: Optional[date] = None, fecha_hasta: Optional[date] = None) -> bytes:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Resumen"

    _write_title(ws1, "Resumen de Inventario", fecha_desde, fecha_hasta)

    headers = ["Indicador", "Valor"]
    rows = [
        ("Total productos", summary.get("total_productos", 0)),
        ("Total unidades", summary.get("total_unidades", 0)),
        ("Valor total", summary.get("valor_total", 0)),
        ("Bajo stock", summary.get("bajo_stock", 0)),
        ("Sin stock", summary.get("sin_stock", 0)),
    ]
    _write_data(ws1, headers, rows)

    for r in range(4, 4 + len(rows)):
        ws1.cell(row=r, column=1).font = BOLD_FONT
        ws1.cell(row=r, column=2).number_format = CURRENCY_FMT

    ws2 = wb.create_sheet("Detalle")
    _write_title(ws2, "Detalle de Inventario", fecha_desde, fecha_hasta)

    detail_headers = ["Producto", "SKU", "Categoría", "Depósito", "Cantidad", "Reservada", "Disponible", "Costo Unit.", "Valor Total"]
    detail_rows = [
        (d["producto"], d["sku"], d["categoria"], d["warehouse"], d["cantidad"], d["reservada"], d["disponible"], d["costo_unitario"], d["valor_total"])
        for d in detail
    ]
    _write_data(ws2, detail_headers, detail_rows)

    _auto_width(ws1)
    _auto_width(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_inventory_rotation(data: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rotación"

    _write_title(ws, "Rotación de Inventario", None, None)

    headers = ["Producto", "SKU", "Ventas 30d", "Stock Actual", "Días Inventario", "Clasificación"]
    rows = [(r["producto"], r["sku"], r["ventas_30d"], r["stock_actual"], r["dias_inventario"], r["clasificacion"]) for r in data]
    _write_data(ws, headers, rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_fiscal_book(data: list, tipo_libro: str, fecha_desde: Optional[date] = None, fecha_hasta: Optional[date] = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Libro {tipo_libro.title()}"

    titulo = f"Libro de {tipo_libro.title()}"
    _write_title(ws, titulo, fecha_desde, fecha_hasta)

    headers = ["Fecha", "Nro Comprobante", "RUC Emisor", "RUC Receptor", "Razón Social", "Cond. IVA", "Base 5%", "Base 10%", "Exento", "IVA 5%", "IVA 10%", "Total"]
    rows = [
        (r["fecha"], r["nro_comprobante"], r["ruc_emisor"], r["ruc_receptor"], r["razon_social"], r["condicion_iva"],
         r["monto_5"], r["monto_10"], r["monto_exento"], r["iva_5"], r["iva_10"], r["total"])
        for r in data
    ]
    _write_data(ws, headers, rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_financial_summary(data: dict, by_day: list, fecha_desde: Optional[date] = None, fecha_hasta: Optional[date] = None) -> bytes:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Resumen"

    _write_title(ws1, "Resumen Financiero", fecha_desde, fecha_hasta)

    headers = ["Indicador", "Valor"]
    rows = [
        ("Ingresos", data.get("ingresos", 0)),
        ("Egresos", data.get("egresos", 0)),
        ("Saldo", data.get("saldo", 0)),
        ("Cuentas por cobrar", data.get("cuentas_por_cobrar", 0)),
        ("Cuentas por pagar", data.get("cuentas_por_pagar", 0)),
        ("Flujo de caja", data.get("flujo_caja", 0)),
    ]
    _write_data(ws1, headers, rows)

    for r in range(4, 4 + len(rows)):
        ws1.cell(row=r, column=1).font = BOLD_FONT
        ws1.cell(row=r, column=2).number_format = CURRENCY_FMT

    ws2 = wb.create_sheet("Diario")
    _write_title(ws2, "Movimiento Diario", fecha_desde, fecha_hasta)

    day_headers = ["Fecha", "Ingresos", "Egresos", "Saldo"]
    day_rows = [(r["fecha"], r["ingresos"], r["egresos"], r["saldo"]) for r in by_day]
    _write_data(ws2, day_headers, day_rows)

    _auto_width(ws1)
    _auto_width(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_fifo_costing(data: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "FIFO"

    _write_title(ws, "Costeo FIFO (Primero en Entrar, Primero en Salir)", None, None)

    headers = ["Producto", "SKU", "Categoría", "Depósito", "Stock Total", "Costo FIFO Unit.", "Valor Total", "Lotes"]
    rows = []
    for item in data:
        lotes_info = "; ".join(
            f"{l['cantidad']}u x Gs.{l['costo_unitario']:,.0f}" for l in item["lotes"][:3]
        )
        if len(item["lotes"]) > 3:
            lotes_info += f" (+{len(item['lotes']) - 3} más)"
        rows.append((
            item["producto"],
            item["sku"],
            item["categoria"],
            item["warehouse"],
            item["total_stock"],
            item["fifo_costo_unitario"],
            item["total_costo"],
            lotes_info,
        ))
    _write_data(ws, headers, rows)

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_lifo_costing(data: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "LIFO"

    _write_title(ws, "Costeo LIFO (Último en Entrar, Primero en Salir)", None, None)

    headers = ["Producto", "SKU", "Categoría", "Depósito", "Stock Total", "Costo LIFO Unit.", "Valor Total", "Lotes"]
    rows = []
    for item in data:
        lotes_info = "; ".join(
            f"{l['cantidad']}u x Gs.{l['costo_unitario']:,.0f}" for l in item["lotes"][:3]
        )
        if len(item["lotes"]) > 3:
            lotes_info += f" (+{len(item['lotes']) - 3} más)"
        rows.append((
            item["producto"],
            item["sku"],
            item["categoria"],
            item["warehouse"],
            item["total_stock"],
            item["lifo_costo_unitario"],
            item["total_costo"],
            lotes_info,
        ))
    _write_data(ws, headers, rows)

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_cost_comparison(data: list) -> bytes:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Comparación"

    _write_title(ws1, "Comparación FIFO vs LIFO vs Promedio Ponderado", None, None)

    headers = ["Producto", "SKU", "Stock", "FIFO", "LIFO", "Prom. Ponderado", "Diferencia", "Dif. %"]
    rows = [
        (
            item["producto"], item["sku"], item["total_stock"],
            item["fifo_costo"], item["lifo_costo"], item["weighted_avg_costo"],
            item["diferencia_fifo_lifo"], f"{item['diferencia_pct']}%",
        )
        for item in data
    ]
    _write_data(ws1, headers, rows)

    ws2 = wb.create_sheet("Detalle FIFO")
    _write_title(ws2, "Detalle de Lotes FIFO", None, None)

    fifo_headers = ["Producto", "Lote", "Cantidad", "Costo Unit.", "Costo Total", "Fecha Ingreso", "Referencia"]
    fifo_rows = []
    for item in data:
        for lote in item.get("lotes_fifo", []):
            fifo_rows.append((
                item["producto"],
                lote.get("lot_id", "")[:8],
                lote["cantidad"],
                lote["costo_unitario"],
                lote["costo_total"],
                lote["fecha_ingreso"],
                lote.get("referencia", ""),
            ))
    _write_data(ws2, fifo_headers, fifo_rows)

    _auto_width(ws1)
    _auto_width(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
