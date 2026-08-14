"""Excel export para Cheques (Cuentas por Pagar) — mismo estilo que
accounts_receivable/export_service.py, reutilizado en vez de reinventado."""

from datetime import date
from decimal import Decimal
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_FONT = Font(name="Inter", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
TITLE_FONT = Font(name="Inter", bold=True, size=14, color="1E40AF")
SUBTITLE_FONT = Font(name="Inter", size=10, color="6B7280")
DATA_FONT = Font(name="Inter", size=10)
BOLD_FONT = Font(name="Inter", bold=True, size=10)
CURRENCY_FMT = "#,##0"
THIN_BORDER = Border(bottom=Side(style="thin", color="E5E7EB"))


def _write_title(ws, titulo, fecha_desde, fecha_hasta, ncols):
    ws.cell(row=1, column=1, value=titulo).font = TITLE_FONT
    periodo = f"Período (fecha emisión): {fecha_desde or 'Inicio'} — {fecha_hasta or 'Actual'}"
    ws.cell(row=2, column=1, value=periodo).font = SUBTITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)


def _write_data(ws, headers, rows, start_row=4):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    for row_idx, row_data in enumerate(rows, start_row + 1):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if isinstance(val, (int, float, Decimal)) and col_idx > 1:
                cell.number_format = CURRENCY_FMT


def _auto_width(ws):
    from openpyxl.utils import get_column_letter
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            idx = getattr(cell, "column", None)
            if not idx:
                continue
            widths[idx] = max(widths.get(idx, 0), len(str(cell.value)))
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(w + 4, 40)


def export_cheques_excel(cheques: list[dict], fecha_desde: Optional[date], fecha_hasta: Optional[date]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cheques"
    _write_title(ws, "Cheques Emitidos", fecha_desde, fecha_hasta, 8)

    headers = ["N° Cheque", "Banco", "Beneficiario", "Monto", "Emisión", "Vencimiento", "Estado", "Días para vencer"]
    rows = []
    total = Decimal("0")
    for c in cheques:
        fecha_pago = c.get("fecha_pago")
        rows.append((
            c.get("numero") or "—",
            c.get("banco_emisor") or "—",
            c.get("supplier_nombre") or c.get("beneficiario") or "—",
            float(c.get("monto") or 0),
            c["fecha_emision"].strftime("%d/%m/%Y") if c.get("fecha_emision") else "",
            fecha_pago.strftime("%d/%m/%Y") if fecha_pago else "",
            c.get("estado") or "—",
            c.get("dias_para_vencer") if c.get("dias_para_vencer") is not None else "",
        ))
        total += Decimal(str(c.get("monto") or 0))
    rows.append(("", "", "TOTAL", float(total), "", "", "", ""))
    _write_data(ws, headers, rows)
    ws.cell(row=4 + len(rows), column=3).font = BOLD_FONT
    ws.cell(row=4 + len(rows), column=4).font = BOLD_FONT
    _auto_width(ws)

    import io
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
