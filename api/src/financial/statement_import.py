"""Parser de extractos bancarios reales (Excel tal cual lo entrega el banco) --
Bancos Fase 6. Convierte en código permanente la lógica de parseo probada a
mano esta sesión sobre los extractos reales de Continental e Interfisa
(agosto 2025 - agosto 2026): hojas por tipo de cuenta (CTE/AHORRO/AHORRO
PIX), fila "Saldo Anterior", columnas DÉBITO/CRÉDITO/SALDO, fila de
totales con el saldo de cierre.

Dos correcciones reales encontradas en esos extractos, aplicadas aquí como
reglas permanentes en vez de parches de una sola vez:
  1. Interfisa: las columnas DÉBITO/CRÉDITO vienen invertidas respecto al
     extracto oficial del banco (confirmado cruzando depósitos de POS
     conocidos). Continental está bien, no se toca.
  2. Fechas con año (o mes y año) mal tipeados en la celda -- se confía en
     el mes/año que declara la carga (elegido explícitamente por quien
     importa, no adivinado del nombre del archivo) por sobre el año de la
     celda, y se recorta el día si hace falta.
"""
import calendar
from datetime import date, datetime
from decimal import Decimal

import openpyxl


def _to_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v if v != 0 else None
    if isinstance(v, str):
        v = v.strip().replace(".", "").replace(",", ".")
        if v in ("", "-"):
            return None
        try:
            n = float(v)
            return n if n != 0 else None
        except ValueError:
            return None
    return None


def _classify_sheet(sheet_name: str, title) -> str | None:
    """Devuelve 'cuenta_corriente' | 'ahorro' | 'ahorro_pix' según el nombre
    de hoja / título, o None si no es una hoja de cuenta reconocible."""
    t = (title if isinstance(title, str) else "").upper()
    s = (sheet_name or "").upper()
    if "PIX" in s or "PIX" in t:
        return "ahorro_pix"
    if "CTE" in s or "CORRIENTE" in s or "CTE" in t or "CORRIENTE" in t:
        return "cuenta_corriente"
    if "AHORRO" in s or "AHORRO" in t:
        return "ahorro"
    return None


def account_tipo_normalizado(tipo: str | None, titular: str | None, numero_cuenta: str | None) -> str:
    """Mapea el tipo/nombre de una BankAccount de InteliMarket al mismo
    vocabulario que usa el extracto ('cuenta_corriente'/'ahorro'/'ahorro_pix'),
    porque tipo en bank_accounts no está normalizado de forma consistente
    (cuentas viejas usan 'cuenta_corriente', el formulario nuevo usa
    'corriente')."""
    blob = f"{titular or ''} {numero_cuenta or ''}".upper()
    if "PIX" in blob:
        return "ahorro_pix"
    t = (tipo or "").lower()
    if "ahorro" in t:
        return "ahorro"
    return "cuenta_corriente"


def parse_statement(file_bytes: bytes, mes: int, anio: int, banco: str, account_tipo: str) -> dict:
    """Parsea un extracto Excel y devuelve las transacciones de la hoja que
    coincide con account_tipo ('cuenta_corriente'/'ahorro'/'ahorro_pix').
    Lanza ValueError si el archivo no tiene ninguna hoja de ese tipo."""
    import io

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    es_interfisa = "INTERFISA" in (banco or "").upper()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        title = rows[0][0] if rows[0] else None
        tipo = _classify_sheet(sheet_name, title)
        if tipo != account_tipo:
            continue

        saldo_anterior = None
        closing_from_totals = None
        txs = []
        for r in rows[1:]:
            fecha, mov, numop, debito, credito, saldo = (r + (None,) * 6)[:6]
            if fecha is None and isinstance(mov, str) and mov.strip() == "Saldo Anterior":
                sn = _to_num(saldo)
                saldo_anterior = sn if sn is not None else 0.0
                continue

            debito_n, credito_n, saldo_n = _to_num(debito), _to_num(credito), _to_num(saldo)
            if es_interfisa:
                debito_n, credito_n = credito_n, debito_n

            if isinstance(fecha, (datetime, date)):
                cell_date = fecha.date() if isinstance(fecha, datetime) else fecha
                if cell_date.month == mes and cell_date.year != anio:
                    fecha_real = date(anio, mes, cell_date.day)
                else:
                    fecha_real = cell_date
                if not (fecha_real.year == anio and fecha_real.month == mes):
                    last_day = calendar.monthrange(anio, mes)[1]
                    fecha_real = date(anio, mes, min(cell_date.day, last_day))

                descripcion = mov or ""
                referencia = str(numop) if numop is not None else None
                if debito_n is not None:
                    txs.append({"fecha": fecha_real, "tipo": "debito", "monto": Decimal(str(debito_n)), "descripcion": descripcion, "referencia": referencia})
                if credito_n is not None:
                    txs.append({"fecha": fecha_real, "tipo": "credito", "monto": Decimal(str(credito_n)), "descripcion": descripcion, "referencia": referencia})
                continue

            if fecha is None and mov is None and (debito_n is not None or credito_n is not None) and saldo_n is not None:
                closing_from_totals = saldo_n

        return {
            "sheet_matched": sheet_name,
            "saldo_anterior": saldo_anterior,
            "closing_from_totals": closing_from_totals,
            "transacciones": txs,
        }

    raise ValueError(f"El archivo no tiene ninguna hoja de tipo '{account_tipo}' (hojas encontradas: {', '.join(wb.sheetnames)})")
