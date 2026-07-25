"""Export service — generates Excel (XLSX) and PDF for delivery analytics."""

import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.units import mm

from api.src.intelientregas.analytics_service import (
    get_profitability_summary,
    get_margins_by_route,
    get_margins_by_driver,
    get_margins_by_vehicle,
    get_delivery_performance_kpi,
    get_margins_by_zone,
    get_business_line_summary,
)
from api.src.intelientregas import service as delivery_service

HEADER_FONT = Font(name="Inter", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
TITLE_FONT = Font(name="Inter", bold=True, size=14, color="1E40AF")
SUBTITLE_FONT = Font(name="Inter", size=10, color="6B7280")
DATA_FONT = Font(name="Inter", size=10)
THIN_BORDER = Border(bottom=Side(style="thin", color="E5E7EB"))


def _style_header(ws, cols):
    for col_idx in range(1, cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


STYLES = getSampleStyleSheet()
PDF_TITLE = ParagraphStyle("Title2", parent=STYLES["Title"], fontSize=16, spaceAfter=12, textColor=colors.HexColor("#1E40AF"))
PDF_H1 = ParagraphStyle("H1", parent=STYLES["Heading2"], fontSize=12, spaceAfter=6, textColor=colors.HexColor("#1E40AF"))
PDF_NORMAL = ParagraphStyle("Normal2", parent=STYLES["Normal"], fontSize=8, spaceAfter=2)
PDF_HEADER = ParagraphStyle("Header2", parent=PDF_NORMAL, fontName="Helvetica-Bold", fontSize=8, textColor=colors.white)


async def export_delivery_excel(db, company_id: str, days: int = 30) -> bytes:
    """Generate complete delivery analytics XLSX."""
    wb = Workbook()

    # ── Sheet 1: Summary ──
    ws1 = wb.active
    ws1.title = "Resumen"
    summary = await get_profitability_summary(db, company_id, days)
    kpi = await get_delivery_performance_kpi(db, company_id, days)

    ws1.cell(row=1, column=1, value=f"Analytics de Entregas — Últimos {days} días").font = TITLE_FONT
    ws1.merge_cells("A1:C1")
    ws1.cell(row=2, column=1, value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = SUBTITLE_FONT
    ws1.merge_cells("A2:C2")

    headers = ["Métrica", "Valor", "Unidad"]
    rows = [
        ("Entregas totales", summary["total_deliveries"], "und"),
        ("Ingresos totales", summary["total_revenue"], "Gs."),
        ("Costo combustible", summary["fuel_cost"], "Gs."),
        ("Costo mantenimiento", summary["maintenance_cost"], "Gs."),
        ("Otros gastos", summary["expense_cost"], "Gs."),
        ("Costo total", summary["total_cost"], "Gs."),
        ("Margen bruto", summary["gross_margin"], "Gs."),
        ("Margen %", summary["margin_pct"], "%"),
        ("Promedio ingreso/entrega", summary["avg_revenue_per_delivery"], "Gs."),
        ("Promedio costo/entrega", summary["avg_cost_per_delivery"], "Gs."),
        ("", "", ""),
        ("Tasa de entrega", kpi["delivery_rate"], "%"),
        ("Tasa de falla", kpi["failed_rate"], "%"),
        ("Total km recorridos", kpi["total_km"], "km"),
        ("Total litros combustible", kpi["total_liters_fuel"], "L"),
        ("Rendimiento", kpi["fuel_efficiency_kmpl"], "km/L"),
    ]
    for i, (h,) in enumerate([(h,) for h in headers], 1):
        ws1.cell(row=3, column=i, value=h)
    _style_header(ws1, 3)
    for idx, (metric, val, unit) in enumerate(rows, 4):
        ws1.cell(row=idx, column=1, value=metric).font = DATA_FONT
        ws1.cell(row=idx, column=2, value=val).font = DATA_FONT
        ws1.cell(row=idx, column=3, value=unit).font = DATA_FONT
        for c in range(1, 4):
            ws1.cell(row=idx, column=c).border = THIN_BORDER
    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 10

    # ── Sheet 2: Routes ──
    ws2 = wb.create_sheet("Rutas")
    routes = await get_margins_by_route(db, company_id, days)
    ws2.cell(row=1, column=1, value="Rentabilidad por Ruta").font = TITLE_FONT
    ws2.merge_cells("A1:G1")
    rh = ["Ruta", "Entregas", "Ingresos", "Costo est.", "Margen", "Margen %", "Distancia (km)"]
    for i, h in enumerate(rh, 1):
        ws2.cell(row=2, column=i, value=h)
    _style_header(ws2, len(rh))
    for idx, r in enumerate(routes, 3):
        vals = [r["route_nombre"], r["deliveries"], r["revenue"], r["estimated_cost"], r["margin"], r["margin_pct"], r["distance_km"]]
        for ci, v in enumerate(vals, 1):
            c = ws2.cell(row=idx, column=ci, value=v)
            c.font = DATA_FONT
            c.border = THIN_BORDER
            if isinstance(v, (int, float)) and ci > 1:
                c.number_format = '#,##0'
    for ci, w in enumerate([30, 10, 15, 15, 15, 10, 15], 1):
        ws2.column_dimensions[chr(64 + ci)].width = w

    # ── Sheet 3: Drivers ──
    ws3 = wb.create_sheet("Repartidores")
    drivers = await get_margins_by_driver(db, company_id, days)
    ws3.cell(row=1, column=1, value="Rentabilidad por Repartidor").font = TITLE_FONT
    ws3.merge_cells("A1:G1")
    dh = ["Repartidor", "Entregas", "Ingresos", "Costo est.", "Margen", "Margen %", "Rating"]
    for i, h in enumerate(dh, 1):
        ws3.cell(row=2, column=i, value=h)
    _style_header(ws3, len(dh))
    for idx, d in enumerate(drivers, 3):
        vals = [d["driver_nombre"], d["deliveries"], d["revenue"], d["estimated_cost"], d["margin"], d["margin_pct"], d["rating"]]
        for ci, v in enumerate(vals, 1):
            c = ws3.cell(row=idx, column=ci, value=v)
            c.font = DATA_FONT
            c.border = THIN_BORDER
            if isinstance(v, (int, float)) and ci > 1:
                c.number_format = '#,##0'
    for ci, w in enumerate([25, 10, 15, 15, 15, 10, 8], 1):
        ws3.column_dimensions[chr(64 + ci)].width = w

    # ── Sheet 4: Vehicles ──
    ws4 = wb.create_sheet("Vehículos")
    vehicles = await get_margins_by_vehicle(db, company_id, days)
    ws4.cell(row=1, column=1, value="Rentabilidad por Vehículo").font = TITLE_FONT
    ws4.merge_cells("A1:I1")
    vh = ["Vehículo", "Tipo", "Entregas", "Ingresos", "Combustible", "Mantenimiento", "Otros gastos", "Costo total", "Margen", "Margen %"]
    for i, h in enumerate(vh, 1):
        ws4.cell(row=2, column=i, value=h)
    _style_header(ws4, len(vh))
    for idx, v in enumerate(vehicles, 3):
        vals = [v["vehicle_label"], v["vehicle_tipo"], v["deliveries"], v["revenue"], v["fuel_cost"], v["maintenance_cost"], v["expense_cost"], v["total_cost"], v["margin"], v["margin_pct"]]
        for ci, v in enumerate(vals, 1):
            c = ws4.cell(row=idx, column=ci, value=v)
            c.font = DATA_FONT
            c.border = THIN_BORDER
            if isinstance(v, (int, float)) and ci > 1:
                c.number_format = '#,##0'
    for ci, w in enumerate([30, 10, 8, 12, 12, 12, 12, 12, 12, 8], 1):
        ws4.column_dimensions[chr(64 + ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def export_delivery_pdf(db, company_id: str, days: int = 30) -> bytes:
    """Generate delivery analytics PDF summary."""
    summary = await get_profitability_summary(db, company_id, days)
    kpi = await get_delivery_performance_kpi(db, company_id, days)
    routes = await get_margins_by_route(db, company_id, days)
    drivers = await get_margins_by_driver(db, company_id, days)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=15*mm, bottomMargin=10*mm, leftMargin=10*mm, rightMargin=10*mm)
    elements = []

    elements.append(Paragraph(f"Analytics de Entregas — Últimos {days} días", PDF_TITLE))
    elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", PDF_NORMAL))
    elements.append(Spacer(1, 6*mm))

    # Summary table
    elements.append(Paragraph("Resumen Financiero", PDF_H1))
    s_data = [
        ["Métrica", "Valor"],
        ["Entregas", str(summary["total_deliveries"])],
        ["Ingresos", f'Gs. {summary["total_revenue"]:,.0f}'],
        ["Costo total", f'Gs. {summary["total_cost"]:,.0f}'],
        ["Margen bruto", f'Gs. {summary["gross_margin"]:,.0f}'],
        ["Margen %", f'{summary["margin_pct"]}%'],
        ["Tasa de entrega", f'{kpi["delivery_rate"]}%'],
        ["Rendimiento", f'{kpi["fuel_efficiency_kmpl"]} km/L'],
    ]
    t = Table(s_data, colWidths=[120, 100])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 6*mm))

    # Top routes table
    if routes:
        elements.append(Paragraph("Top Rutas por Ingreso", PDF_H1))
        r_data = [["Ruta", "Entregas", "Ingresos", "Margen %"]] + [
            [r["route_nombre"][:25], str(r["deliveries"]), f'Gs. {r["revenue"]:,.0f}', f'{r["margin_pct"]}%']
            for r in routes[:10]
        ]
        t2 = Table(r_data, colWidths=[130, 60, 100, 60])
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ]))
        elements.append(t2)
        elements.append(Spacer(1, 4*mm))

    # Top drivers table
    if drivers:
        elements.append(Paragraph("Top Repartidores por Ingreso", PDF_H1))
        d_data = [["Repartidor", "Entregas", "Ingresos", "Margen %", "Rating"]] + [
            [d["driver_nombre"][:20], str(d["deliveries"]), f'Gs. {d["revenue"]:,.0f}', f'{d["margin_pct"]}%', str(d["rating"])]
            for d in drivers[:10]
        ]
        t3 = Table(d_data, colWidths=[100, 60, 100, 60, 40])
        t3.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ]))
        elements.append(t3)

    doc.build(elements)
    buf.seek(0)
    return buf.read()
