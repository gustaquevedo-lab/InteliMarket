"""Excel export del Kardex -- mismo estilo que accounts_receivable/export_service.py
(reutilizado, no reinventado): titulo + periodo en las primeras filas, encabezado
de tabla con fondo azul, auto-ancho de columnas."""

from datetime import date
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(name="Inter", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
TITLE_FONT = Font(name="Inter", bold=True, size=14, color="1E40AF")
SUBTITLE_FONT = Font(name="Inter", size=10, color="6B7280")
DATA_FONT = Font(name="Inter", size=10)
BOLD_FONT = Font(name="Inter", bold=True, size=10)
THIN_BORDER = Border(bottom=Side(style="thin", color="E5E7EB"))
POSITIVE_FONT = Font(name="Inter", size=10, color="059669")
NEGATIVE_FONT = Font(name="Inter", size=10, color="DC2626")


def export_kardex_excel(movements: list[dict], fecha_desde: Optional[date], fecha_hasta: Optional[date]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Kardex"

    ncols = 9
    ws.cell(row=1, column=1, value="Libro Kardex & Trazabilidad de Inventario").font = TITLE_FONT
    periodo = f"Período: {fecha_desde or 'Inicio'} — {fecha_hasta or 'Actual'}  ·  {len(movements)} movimientos"
    ws.cell(row=2, column=1, value=periodo).font = SUBTITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)

    headers = ["Fecha & Hora", "Tipo", "Producto", "SKU", "Depósito", "Cantidad", "Saldo Acumulado", "Usuario", "Motivo / Documento"]
    start_row = 4
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for row_idx, m in enumerate(movements, start_row + 1):
        cantidad = float(m.get("cantidad") or 0)
        vals = [
            m["created_at"].strftime("%d/%m/%Y %H:%M") if hasattr(m.get("created_at"), "strftime") else str(m.get("created_at") or ""),
            m.get("tipo") or "",
            m.get("product_nombre") or "Producto",
            m.get("product_sku") or "",
            m.get("warehouse_nombre") or "Depósito Central",
            cantidad,
            float(m.get("saldo_acumulado") or 0),
            m.get("user_nombre") or "—",
            m.get("motivo") or "Movimiento operativo",
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if col_idx == 6:
                cell.font = POSITIVE_FONT if cantidad >= 0 else NEGATIVE_FONT
                cell.number_format = "+#,##0;-#,##0"
            elif col_idx == 7:
                cell.font = BOLD_FONT
                cell.number_format = "#,##0"
            else:
                cell.font = DATA_FONT

    widths: dict[int, int] = {}
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(w + 4, 45)

    import io
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
