"""Excel exports para Cuentas por Cobrar — mismo estilo que reports/export_service.py
(reutilizado, no reinventado): titulo + periodo en las primeras filas, encabezado
de tabla con fondo azul, auto-ancho de columnas."""

from datetime import date
from decimal import Decimal
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_FONT = Font(name="Inter", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
TITLE_FONT = Font(name="Inter", bold=True, size=14, color="1E40AF")
SUBTITLE_FONT = Font(name="Inter", size=10, color="6B7280")
DATA_FONT = Font(name="Inter", size=10)
BOLD_FONT = Font(name="Inter", bold=True, size=10)
CURRENCY_FMT = "#,##0"
THIN_BORDER = Border(bottom=Side(style="thin", color="E5E7EB"))


def _write_title(ws, titulo, fecha_desde, fecha_hasta, ncols):
    ws.cell(row=1, column=1, value=titulo).font = TITLE_FONT
    periodo = f"Período: {fecha_desde or 'Inicio'} — {fecha_hasta or 'Actual'}"
    ws.cell(row=2, column=1, value=periodo).font = SUBTITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)


def _write_data(ws, headers, rows, start_row=4):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    for row_idx, row_data in enumerate(rows, start_row + 1):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if isinstance(val, (int, float, Decimal)) and col_idx > 1:
                cell.number_format = CURRENCY_FMT


def _auto_width(ws):
    from openpyxl.utils import get_column_letter
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            idx = getattr(cell, "column", None)
            if not idx:
                continue
            widths[idx] = max(widths.get(idx, 0), len(str(cell.value)))
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(w + 4, 40)


def export_aging_excel(aging: dict, fecha_desde: Optional[date], fecha_hasta: Optional[date]) -> bytes:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Por Cliente"
    _write_title(ws1, "Antigüedad de Saldos (Aging)", fecha_desde, fecha_hasta, 8)
    headers = ["Cliente", "Documentos", "Al día", "1-30 días", "31-60 días", "61-90 días", "+90 días", "Saldo Total"]
    rows = [
        (
            c["customer_name"], c["total_documentos"], float(c["current"]), float(c["days_1_30"]),
            float(c["days_31_60"]), float(c["days_61_90"]), float(c["days_91_plus"]), float(c["saldo_total"]),
        )
        for c in aging.get("por_clientes", [])
    ]
    rows.append(("TOTAL", aging.get("cantidad_documentos", 0), float(aging.get("current", 0)), float(aging.get("days_1_30", 0)), float(aging.get("days_31_60", 0)), float(aging.get("days_61_90", 0)), float(aging.get("days_91_plus", 0)), float(aging.get("total_pendiente", 0))))
    _write_data(ws1, headers, rows)
    for c in range(1, 9):
        ws1.cell(row=4 + len(rows), column=c).font = BOLD_FONT
    _auto_width(ws1)

    ws2 = wb.create_sheet("Documentos")
    _write_title(ws2, "Documentos pendientes", fecha_desde, fecha_hasta, 7)
    headers2 = ["N° Documento", "Cliente", "Emisión", "Vencimiento", "Monto Original", "Saldo Pendiente", "Días Mora"]
    rows2 = [
        (
            d["numero_documento"], d["customer_name"],
            d["fecha_emision"].strftime("%d/%m/%Y") if d.get("fecha_emision") else "",
            d["fecha_vencimiento"].strftime("%d/%m/%Y") if d.get("fecha_vencimiento") else "",
            float(d["monto_original"]), float(d["saldo_pendiente"]), d.get("dias_mora") or 0,
        )
        for d in aging.get("documentos", [])
    ]
    _write_data(ws2, headers2, rows2)
    _auto_width(ws2)

    import io
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_cobranzas_excel(payments: list[dict], fecha_desde: Optional[date], fecha_hasta: Optional[date]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cobranzas"
    _write_title(ws, "Cobranzas del período", fecha_desde, fecha_hasta, 6)

    headers = ["Fecha", "Cliente", "Forma de Pago", "Referencia", "Facturas cubiertas", "Monto"]
    rows = []
    total = Decimal("0")
    for p in payments:
        docs = ", ".join(a.get("numero_documento") or "" for a in p.get("allocations", []))
        rows.append((
            p["fecha"].strftime("%d/%m/%Y") if hasattr(p["fecha"], "strftime") else str(p["fecha"]),
            p.get("customer_name") or "—", p.get("forma_pago") or "—", p.get("referencia") or "—",
            docs, float(p["monto_total"]),
        ))
        total += Decimal(str(p["monto_total"]))
    rows.append(("", "", "", "", "TOTAL", float(total)))
    _write_data(ws, headers, rows)
    ws.cell(row=4 + len(rows), column=5).font = BOLD_FONT
    ws.cell(row=4 + len(rows), column=6).font = BOLD_FONT
    _auto_width(ws)

    import io
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
