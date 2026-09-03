from typing import Optional
from datetime import date, datetime, timedelta
from uuid import UUID
from io import BytesIO

from sqlalchemy import select, func as sa_func
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from api.src.sales.models import Sale, SaleItem
from api.src.products.models import Product, ProductCategory
from api.src.inventory.models import Stock
from api.src.supermer.models import ProductionOrder, ProductionRecipe, WasteLog, MarkdownLog
from api.src.accounts_receivable.service import get_receivable_summary
from api.src.financial.service import get_ap_dashboard
from api.src.financial.models import SupplierInvoice
from api.src.petty_cash.models import Expense, ExpenseCategory


AREA_LABELS = {
    "carniceria": "Carnicería",
    "panaderia": "Panadería",
    "rotiseria": "Rotisería",
    "pre_pack": "Pre-pack",
    "verduleria": "Verdulería",
    "pasteleria": "Pastelería",
    "lacteos": "Lácteos",
    "otros": "Otros",
}


def _apply_date_range(q, col, desde, hasta):
    if desde:
        q = q.where(col >= desde)
    if hasta:
        q = q.where(col <= hasta + timedelta(days=1))
    return q


# Mismo criterio que usa caja/service.py para el arqueo real de caja --
# antes este archivo solo contaba "confirmado", dejando afuera ventas en
# "completada"/"completado"/"pagado" que si entraban en el cierre de caja
# real, asi que el resumen gerencial mostraba menos ventas que el arqueo.
VENTA_CONTADA_ESTADOS = ("confirmado", "completada", "completado", "pagado")


async def get_dashboard(
    db: AsyncSession, company_id: str,
    desde: Optional[date] = None, hasta: Optional[date] = None,
) -> dict:
    now = datetime.utcnow()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = today_start - timedelta(days=today_start.weekday())
    month_start = today_start.replace(day=1)

    base_filter = lambda q: q.where(Sale.company_id == company_id, Sale.estado.in_(VENTA_CONTADA_ESTADOS))

    r = await db.execute(
        _apply_date_range(select(sa_func.coalesce(sa_func.sum(Sale.total), 0)),
                          Sale.fecha, today_start, today_start)
        .where(Sale.company_id == company_id, Sale.estado.in_(VENTA_CONTADA_ESTADOS))
    )
    ventas_hoy = float(r.scalar() or 0)

    r = await db.execute(
        _apply_date_range(select(sa_func.coalesce(sa_func.sum(Sale.total), 0)),
                          Sale.fecha, week_start, now)
        .where(Sale.company_id == company_id, Sale.estado.in_(VENTA_CONTADA_ESTADOS))
    )
    ventas_semana = float(r.scalar() or 0)

    r = await db.execute(
        _apply_date_range(select(sa_func.coalesce(sa_func.sum(Sale.total), 0)),
                          Sale.fecha, month_start, now)
        .where(Sale.company_id == company_id, Sale.estado.in_(VENTA_CONTADA_ESTADOS))
    )
    ventas_mes = float(r.scalar() or 0)

    period_q = select(
        sa_func.coalesce(sa_func.sum(Sale.total), 0),
        sa_func.count(Sale.id.distinct()),
        sa_func.count(Sale.customer_id.distinct()),
    ).where(Sale.company_id == company_id, Sale.estado.in_(VENTA_CONTADA_ESTADOS))
    period_q = _apply_date_range(period_q, Sale.fecha, desde or month_start, hasta or now)
    r = await db.execute(period_q)
    row = r.one()
    total_ventas = float(row[0] or 0)
    total_transacciones = row[1] or 0
    total_clientes = row[2] or 0
    ticket_promedio = round(total_ventas / max(total_transacciones, 1), 0)

    items_q = select(sa_func.coalesce(sa_func.sum(SaleItem.cantidad), 0))\
        .join(Sale, SaleItem.sale_id == Sale.id)\
        .where(Sale.company_id == company_id, Sale.estado.in_(VENTA_CONTADA_ESTADOS))
    items_q = _apply_date_range(items_q, Sale.fecha, desde or month_start, hasta or now)
    r = await db.execute(items_q)
    productos_vendidos = int(r.scalar() or 0)

    top_q = select(
        SaleItem.product_id,
        sa_func.sum(SaleItem.cantidad).label("cantidad"),
        sa_func.sum(SaleItem.total).label("total"),
        sa_func.avg(SaleItem.costo_unitario).label("costo_prom"),
    ).join(Sale, SaleItem.sale_id == Sale.id)\
     .where(Sale.company_id == company_id, Sale.estado.in_(VENTA_CONTADA_ESTADOS))
    top_q = _apply_date_range(top_q, Sale.fecha, desde or month_start, hasta or now)
    top_q = top_q.group_by(SaleItem.product_id).order_by(sa_func.sum(SaleItem.total).desc()).limit(10)
    r = await db.execute(top_q)
    top_rows = r.all()

    top_productos = []
    for tr in top_rows:
        prod_r = await db.execute(select(Product).where(Product.id == tr.product_id))
        prod = prod_r.scalar_one_or_none()
        cat_nombre = None
        if prod and prod.categoria_id:
            cat_r = await db.execute(select(ProductCategory.nombre).where(ProductCategory.id == prod.categoria_id))
            cat_nombre = cat_r.scalar_one_or_none()
        cant = float(tr.cantidad or 0)
        tot = float(tr.total or 0)
        costo = float(tr.costo_prom or 0) * cant
        margen = round(((tot - costo) / max(tot, 1)) * 100, 1)
        participacion = round((tot / max(total_ventas, 1)) * 100, 1)
        top_productos.append({
            "producto_id": str(tr.product_id),
            "producto_nombre": prod.nombre if prod else "N/A",
            "categoria": cat_nombre,
            "cantidad_vendida": cant,
            "total_ventas": tot,
            "margen": margen,
            "rotacion_dias": None,
            "participacion_porcentaje": participacion,
        })

    hour_q = select(
        sa_func.extract("hour", Sale.fecha).label("hora"),
        sa_func.coalesce(sa_func.sum(Sale.total), 0).label("total"),
        sa_func.count(Sale.id).label("count"),
    ).where(Sale.company_id == company_id, Sale.estado.in_(VENTA_CONTADA_ESTADOS))
    hour_q = _apply_date_range(hour_q, Sale.fecha, desde or month_start, hasta or now)
    hour_q = hour_q.group_by(sa_func.extract("hour", Sale.fecha)).order_by(sa_func.extract("hour", Sale.fecha))
    r = await db.execute(hour_q)
    hour_rows = r.all()

    hour_map = {int(h.hora): (float(h.total), h.count) for h in hour_rows}
    ventas_por_hora = []
    for hr in range(24):
        total_h, cnt = hour_map.get(hr, (0, 0))
        ventas_por_hora.append({
            "hora": hr,
            "total_ventas": total_h,
            "cantidad_transacciones": cnt,
            "ticket_promedio": round(total_h / max(cnt, 1), 0),
        })

    deptos = await get_depto_pyl(db, company_id, desde, hasta)
    margen_deptos = [d["margen_porcentaje"] for d in deptos if d["ventas"] > 0]

    return {
        "ventas_hoy": ventas_hoy,
        "ventas_semana": ventas_semana,
        "ventas_mes": ventas_mes,
        "margen_promedio": round(sum(margen_deptos) / max(len(margen_deptos), 1), 1) if margen_deptos else 0,
        "ticket_promedio": ticket_promedio,
        "clientes_atendidos": total_clientes,
        "productos_vendidos": productos_vendidos,
        "top_productos": top_productos,
        "ventas_por_hora": ventas_por_hora,
        "deptos": deptos,
    }


async def get_depto_pyl(
    db: AsyncSession, company_id: str,
    desde: Optional[date] = None, hasta: Optional[date] = None,
) -> list[dict]:
    recipe_q = await db.execute(
        select(ProductionRecipe.area, ProductionRecipe.producto_terminado_id)
        .where(ProductionRecipe.activa == True)
    )
    area_products: dict[str, list] = {}
    for row in recipe_q.all():
        area = row.area.value if hasattr(row.area, "value") else str(row.area)
        area_products.setdefault(area, []).append(str(row.producto_terminado_id))

    result = []
    for area, product_ids in area_products.items():
        if not product_ids:
            continue

        pid_uuids = [UUID(pid) for pid in product_ids]

        sales_q = select(
            sa_func.coalesce(sa_func.sum(SaleItem.total), 0),
            sa_func.coalesce(sa_func.sum(SaleItem.costo_unitario * SaleItem.cantidad), 0),
        ).join(Sale, SaleItem.sale_id == Sale.id).where(
            Sale.company_id == company_id,
            Sale.estado.in_(VENTA_CONTADA_ESTADOS),
            SaleItem.product_id.in_(pid_uuids),
        )
        sales_q = _apply_date_range(sales_q, Sale.fecha, desde, hasta)
        r = await db.execute(sales_q)
        srow = r.one()
        ventas = float(srow[0] or 0)
        costo = float(srow[1] or 0)

        waste_q = select(
            sa_func.coalesce(sa_func.sum(WasteLog.cantidad * WasteLog.costo_unitario), 0),
        ).where(WasteLog.company_id == company_id, WasteLog.area == area)
        waste_q = _apply_date_range(waste_q, WasteLog.fecha, desde, hasta)
        r = await db.execute(waste_q)
        merma_costo = float(r.scalar() or 0)

        md_q = select(sa_func.count(MarkdownLog.id)).where(
            MarkdownLog.company_id == company_id,
            MarkdownLog.activo == True,
            MarkdownLog.producto_id.in_(pid_uuids),
        )
        r = await db.execute(md_q)
        markdowns = r.scalar() or 0

        margen = ventas - costo
        result.append({
            "depto": AREA_LABELS.get(area, area.capitalize()),
            "ventas": round(ventas, 0),
            "costo_ventas": round(costo, 0),
            "margen_bruto": round(margen, 0),
            "margen_porcentaje": round((margen / max(ventas, 1)) * 100, 1) if ventas > 0 else 0,
            "merma_total": round(merma_costo, 0),
            "merma_porcentaje": round((merma_costo / max(ventas, 1)) * 100, 1) if ventas > 0 else 0,
            "markdowns_activos": markdowns,
        })

    all_area_pids = set()
    for pids in area_products.values():
        all_area_pids.update(pids)

    if all_area_pids:
        gral_q = select(
            sa_func.coalesce(sa_func.sum(SaleItem.total), 0),
            sa_func.coalesce(sa_func.sum(SaleItem.costo_unitario * SaleItem.cantidad), 0),
        ).join(Sale, SaleItem.sale_id == Sale.id).where(
            Sale.company_id == company_id,
            Sale.estado.in_(VENTA_CONTADA_ESTADOS),
            ~SaleItem.product_id.in_([UUID(pid) for pid in all_area_pids]),
        )
        gral_q = _apply_date_range(gral_q, Sale.fecha, desde, hasta)
        r = await db.execute(gral_q)
        grow = r.one()
        g_ventas = float(grow[0] or 0)
        g_costo = float(grow[1] or 0)
        if g_ventas > 0:
            result.append({
                "depto": "General",
                "ventas": round(g_ventas, 0),
                "costo_ventas": round(g_costo, 0),
                "margen_bruto": round(g_ventas - g_costo, 0),
                "margen_porcentaje": round(((g_ventas - g_costo) / max(g_ventas, 1)) * 100, 1),
                "merma_total": 0,
                "merma_porcentaje": 0,
                "markdowns_activos": 0,
            })

    if not result:
        r = await db.execute(
            select(
                sa_func.coalesce(sa_func.sum(SaleItem.total), 0),
                sa_func.coalesce(sa_func.sum(SaleItem.costo_unitario * SaleItem.cantidad), 0),
            )
            .join(Sale, SaleItem.sale_id == Sale.id)
            .where(Sale.company_id == company_id, Sale.estado.in_(VENTA_CONTADA_ESTADOS))
        )
        row = r.one()
        ventas = float(row[0] or 0)
        costo = float(row[1] or 0)
        result.append({
            "depto": "General",
            "ventas": round(ventas, 0),
            "costo_ventas": round(costo, 0),
            "margen_bruto": round(ventas - costo, 0),
            "margen_porcentaje": round(((ventas - costo) / max(ventas, 1)) * 100, 1),
            "merma_total": 0,
            "merma_porcentaje": 0,
            "markdowns_activos": 0,
        })

    return result


async def get_ranking(
    db: AsyncSession, company_id: str,
    desde: Optional[date] = None, hasta: Optional[date] = None,
    limit: int = 20,
) -> list[dict]:
    total_q = select(sa_func.coalesce(sa_func.sum(SaleItem.total), 0))\
        .join(Sale, SaleItem.sale_id == Sale.id)\
        .where(Sale.company_id == company_id, Sale.estado.in_(VENTA_CONTADA_ESTADOS))
    total_q = _apply_date_range(total_q, Sale.fecha, desde, hasta)
    r = await db.execute(total_q)
    total_ventas = float(r.scalar() or 1)

    q = select(
        SaleItem.product_id,
        sa_func.sum(SaleItem.cantidad).label("cantidad"),
        sa_func.sum(SaleItem.total).label("total"),
        sa_func.avg(SaleItem.costo_unitario).label("costo_prom"),
    ).join(Sale, SaleItem.sale_id == Sale.id)\
     .where(Sale.company_id == company_id, Sale.estado.in_(VENTA_CONTADA_ESTADOS))
    q = _apply_date_range(q, Sale.fecha, desde, hasta)
    q = q.group_by(SaleItem.product_id).order_by(sa_func.sum(SaleItem.total).desc()).limit(limit)
    r = await db.execute(q)
    rows = r.all()

    num_days = 30
    if desde and hasta:
        num_days = max((hasta - desde).days, 1)

    result = []
    for row in rows:
        prod_r = await db.execute(select(Product).where(Product.id == row.product_id))
        prod = prod_r.scalar_one_or_none()
        cat_nombre = None
        if prod and prod.categoria_id:
            cat_r = await db.execute(select(ProductCategory.nombre).where(ProductCategory.id == prod.categoria_id))
            cat_nombre = cat_r.scalar_one_or_none()
        cant = float(row.cantidad or 0)
        tot = float(row.total or 0)
        costo = float(row.costo_prom or 0) * cant
        margen = round(((tot - costo) / max(tot, 1)) * 100, 1)
        participacion = round((tot / total_ventas) * 100, 2)

        stock_q = await db.execute(
            select(sa_func.coalesce(sa_func.sum(Stock.cantidad), 0))
            .where(Stock.product_id == row.product_id)
        )
        stock = float(stock_q.scalar() or 0)
        avg_daily = cant / num_days
        rotacion = round(stock / max(avg_daily, 0.01), 1) if avg_daily > 0 else None

        result.append({
            "producto_id": str(row.product_id),
            "producto_nombre": prod.nombre if prod else "N/A",
            "categoria": cat_nombre,
            "cantidad_vendida": cant,
            "total_ventas": tot,
            "margen": margen,
            "rotacion_dias": rotacion,
            "participacion_porcentaje": participacion,
        })

    return result


async def get_alertas_negocio(
    db: AsyncSession, company_id: str,
    margen_umbral: float = 15.0,
) -> dict:
    """Alertas de margen real y días de cobro vs. pago, sobre datos reales de los últimos 30 días."""
    today = date.today()
    since_30 = today - timedelta(days=30)

    q = select(
        SaleItem.product_id,
        sa_func.sum(SaleItem.cantidad).label("cantidad"),
        sa_func.sum(SaleItem.total).label("total"),
        sa_func.avg(SaleItem.costo_unitario).label("costo_prom"),
    ).join(Sale, SaleItem.sale_id == Sale.id).where(
        Sale.company_id == company_id,
        Sale.estado.in_(VENTA_CONTADA_ESTADOS),
        Sale.fecha >= since_30,
    ).group_by(SaleItem.product_id).having(sa_func.sum(SaleItem.cantidad) >= 3)
    r = await db.execute(q)
    rows = r.all()

    margen_bajo = []
    for row in rows:
        cant = float(row.cantidad or 0)
        tot = float(row.total or 0)
        costo = float(row.costo_prom or 0) * cant
        margen_pct = round(((tot - costo) / max(tot, 1)) * 100, 1)
        if tot > 0 and margen_pct < margen_umbral:
            margen_bajo.append((margen_pct, row.product_id, cant, tot, margen_pct))
    margen_bajo.sort(key=lambda x: x[0])
    margen_bajo = margen_bajo[:10]

    margen_bajo_out = []
    for _, product_id, cant, tot, margen_pct in margen_bajo:
        prod_r = await db.execute(select(Product.nombre).where(Product.id == product_id))
        nombre = prod_r.scalar_one_or_none() or "N/A"
        margen_bajo_out.append({
            "producto_id": str(product_id),
            "producto_nombre": nombre,
            "cantidad_vendida_30d": cant,
            "total_ventas_30d": tot,
            "margen_porcentaje": margen_pct,
        })

    ar = await get_receivable_summary(db, company_id)
    ap = await get_ap_dashboard(db, company_id)

    ventas_r = await db.execute(
        select(sa_func.coalesce(sa_func.sum(Sale.total), 0))
        .where(Sale.company_id == company_id, Sale.estado.in_(VENTA_CONTADA_ESTADOS), Sale.fecha >= since_30)
    )
    ventas_30d = float(ventas_r.scalar() or 0)

    compras_r = await db.execute(
        select(sa_func.coalesce(sa_func.sum(SupplierInvoice.total), 0))
        .where(SupplierInvoice.company_id == UUID(company_id), SupplierInvoice.fecha_emision >= since_30)
    )
    compras_30d = float(compras_r.scalar() or 0)

    monto_pendiente_ar = float(ar.get("total_pendiente") or 0)
    monto_pendiente_ap = float(ap.get("total_pendiente") or 0)
    dias_cobro = round(monto_pendiente_ar / (ventas_30d / 30), 1) if ventas_30d > 0 else None
    dias_pago = round(monto_pendiente_ap / (compras_30d / 30), 1) if compras_30d > 0 else None

    return {
        "margen_bajo": margen_bajo_out,
        "margen_umbral": margen_umbral,
        "cxc_vencidas": {
            "cantidad": int(ar.get("vencidos") or 0),
            "monto": float(ar.get("monto_vencido") or 0),
            "total_pendiente": monto_pendiente_ar,
        },
        "cxp_vencidas": {
            "cantidad": int(ap.get("facturas_vencidas") or 0),
            "monto": float(ap.get("total_vencido") or 0),
            "total_pendiente": monto_pendiente_ap,
        },
        "dias_cobro_promedio": dias_cobro,
        "dias_pago_promedio": dias_pago,
    }


async def get_pnl_data(
    db: AsyncSession, company_id: str,
    desde: Optional[date] = None, hasta: Optional[date] = None,
) -> dict:
    """Estado de Resultados real: ventas y costo de mercaderia de sales/sale_items,
    gastos operativos reales de petty_cash.expenses (agrupados por categoria)."""
    today = date.today()
    desde = desde or today.replace(day=1)
    hasta = hasta or today

    r = await db.execute(
        select(sa_func.coalesce(sa_func.sum(Sale.total), 0))
        .where(Sale.company_id == company_id, Sale.estado.in_(VENTA_CONTADA_ESTADOS),
               Sale.fecha >= desde, Sale.fecha <= hasta + timedelta(days=1))
    )
    total_ventas = float(r.scalar() or 0)

    r = await db.execute(
        select(sa_func.coalesce(sa_func.sum(SaleItem.costo_unitario * SaleItem.cantidad), 0))
        .join(Sale, SaleItem.sale_id == Sale.id)
        .where(Sale.company_id == company_id, Sale.estado.in_(VENTA_CONTADA_ESTADOS),
               Sale.fecha >= desde, Sale.fecha <= hasta + timedelta(days=1))
    )
    total_costo = float(r.scalar() or 0)

    r = await db.execute(
        select(
            sa_func.coalesce(ExpenseCategory.nombre, "Sin categoría"),
            sa_func.sum(Expense.monto),
        )
        .select_from(Expense)
        .outerjoin(ExpenseCategory, ExpenseCategory.id == Expense.category_id)
        .where(Expense.company_id == company_id, Expense.fecha_gasto >= desde, Expense.fecha_gasto <= hasta)
        .group_by(ExpenseCategory.nombre)
        .order_by(sa_func.sum(Expense.monto).desc())
    )
    gastos = [{"nombre": nombre, "monto": float(monto or 0)} for nombre, monto in r.all()]
    total_gastos = sum(g["monto"] for g in gastos)

    resultado_bruto = total_ventas - total_costo
    resultado_neto = resultado_bruto - total_gastos

    return {
        "periodo": f"{desde.strftime('%d/%m/%Y')} — {hasta.strftime('%d/%m/%Y')}",
        "ingresos": [{"nombre": "Ventas", "monto": total_ventas}],
        "total_ingresos": total_ventas,
        "costos": [{"nombre": "Costo de Mercadería Vendida", "monto": total_costo}],
        "total_costos": total_costo,
        "resultado_bruto": resultado_bruto,
        "gastos": gastos,
        "total_gastos": total_gastos,
        "resultado_neto": resultado_neto,
    }


async def export_excel(
    db: AsyncSession, company_id: str, report_type: str,
    desde: Optional[date] = None, hasta: Optional[date] = None,
) -> bytes:
    wb = Workbook()

    if report_type == "dashboard":
        data = await get_dashboard(db, company_id, desde, hasta)
        ws = wb.active
        ws.title = "Dashboard"
        _write_title(ws, "Dashboard Gerencial", desde, hasta)

        kpis = [
            ("Ventas Hoy", data["ventas_hoy"]),
            ("Ventas Semana", data["ventas_semana"]),
            ("Ventas Mes", data["ventas_mes"]),
            ("Margen Promedio", f"{data['margen_promedio']}%"),
            ("Ticket Promedio", data["ticket_promedio"]),
            ("Clientes Atendidos", data["clientes_atendidos"]),
            ("Productos Vendidos", data["productos_vendidos"]),
        ]
        _write_table(ws, ["Indicador", "Valor"], kpis)

        ws2 = wb.create_sheet("Top Productos")
        _write_title(ws2, "Top 10 Productos", desde, hasta)
        _write_table(ws2, ["Producto", "Cantidad", "Ventas", "Margen %", "Participación"],
                     [[p["producto_nombre"], p["cantidad_vendida"], p["total_ventas"],
                       f"{p['margen']}%", f"{p['participacion_porcentaje']}%"] for p in data["top_productos"]])

        ws3 = wb.create_sheet("Ventas por Hora")
        _write_title(ws3, "Ventas por Hora", desde, hasta)
        _write_table(ws3, ["Hora", "Ventas", "Transacciones", "Ticket Promedio"],
                     [[f"{h['hora']:02d}:00", h["total_ventas"], h["cantidad_transacciones"], h["ticket_promedio"]]
                      for h in data["ventas_por_hora"]])

    elif report_type == "deptos":
        data = await get_depto_pyl(db, company_id, desde, hasta)
        ws = wb.active
        ws.title = "P&L Departamentos"
        _write_title(ws, "P&L por Departamento", desde, hasta)
        _write_table(ws, ["Departamento", "Ventas", "Costo Ventas", "Margen Bruto", "Margen %", "Merma", "Merma %", "Markdowns"],
                     [[d["depto"], d["ventas"], d["costo_ventas"], d["margen_bruto"],
                       f"{d['margen_porcentaje']}%", d["merma_total"], f"{d['merma_porcentaje']}%",
                       d["markdowns_activos"]] for d in data])

    elif report_type == "ranking":
        data = await get_ranking(db, company_id, desde, hasta)
        ws = wb.active
        ws.title = "Ranking Productos"
        _write_title(ws, "Ranking de Productos", desde, hasta)
        _write_table(ws, ["Producto", "Categoría", "Cantidad", "Ventas", "Margen %", "Rotación (días)", "Participación"],
                     [[p["producto_nombre"], p["categoria"] or "", p["cantidad_vendida"], p["total_ventas"],
                       f"{p['margen']}%", p["rotacion_dias"] or "—", f"{p['participacion_porcentaje']}%"]
                      for p in data])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


HEADER_FILL = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
HEADER_FONT = Font(name="Inter", bold=True, size=11, color="FFFFFF")
TITLE_FONT = Font(name="Inter", bold=True, size=14, color="1E40AF")
SUBTITLE_FONT = Font(name="Inter", size=10, color="6B7280")
DATA_FONT = Font(name="Inter", size=10)
THIN_BORDER = Border(bottom=Side(style="thin", color="E5E7EB"))


def _write_title(ws, titulo, desde, hasta):
    ws.cell(row=1, column=1, value=titulo).font = TITLE_FONT
    periodo = f"Período: {desde or 'Inicio'} — {hasta or 'Actual'}" if desde or hasta else "Todos los períodos"
    ws.cell(row=2, column=1, value=periodo).font = SUBTITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)


def _write_table(ws, headers, rows, start_row=4):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    for row_idx, row_data in enumerate(rows, start_row + 1):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
